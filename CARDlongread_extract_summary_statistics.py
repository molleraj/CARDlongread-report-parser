#!/usr/bin/env python3

import pandas as pd
import numpy as np
import seaborn as sb
import matplotlib.pyplot as plt
import statistics
import argparse
import dataclasses
# excel export with openpyxl
import openpyxl
# and xlsxwriter with pandas, though that could be modified to use openpyxl as well
import xlsxwriter
# for image saving
from io import BytesIO

# get summary statistics (min, max, range, mean, median, mode, and standard deviation for N50, sequence output, and flow cells per sample)
def get_summary_statistics(column):
    # define summary statistics class
    @dataclasses.dataclass
    class summary_statistics:
        total : int = 0
        min : float = 0
        max : float = 0
        range : float = 0
        mean : float = 0
        median : float = 0
        mode : float = 0
        stdev : float = 0
	# take single column of data frame as input
	# fill summary statistics class with each statistic
    # note that the line below should be used to give the total run count (across experiments and samples). Not useful with unique experiments, samples, etc.
    # as in flow cells per unique sample
    summary_statistics.total = len(column)
    summary_statistics.min = min(column)
    summary_statistics.max = max(column)
    summary_statistics.range = summary_statistics.max - summary_statistics.min
    summary_statistics.mean = statistics.mean(column)
    summary_statistics.median = statistics.median(column)
    summary_statistics.mode = statistics.mode(column)
    summary_statistics.stdev = statistics.stdev(column)
    # return data structure with each summary statistic as an attribute
    return summary_statistics
    
# get output per flow cell in two column list
def get_output_per_flow_cell(flow_cell_IDs, output, topup):
    # make data frame of flow cell IDs and output
    flow_cells_to_output = pd.concat([flow_cell_IDs, output, topup], axis=1, join='inner')
    # find unique flow cells
    unique_flow_cells = np.unique(flow_cell_IDs.astype(str))
    # create output_per_flow_cell_df data frame
    output_per_flow_cell_df = pd.DataFrame({'Flow Cell ID' : unique_flow_cells}, index=unique_flow_cells, columns=['Flow Cell ID','Flow cell output (Gb)','Top up'])
    # find flow cells and total output per unique experiment by iterating through unique experiment names
    for i in unique_flow_cells:
        unique_outputs_per_flow_cell = pd.unique(flow_cells_to_output[flow_cells_to_output['Flow Cell ID'] == i]['Data output (Gb)'])
        unique_topup_per_flow_cell = pd.unique(flow_cells_to_output[flow_cells_to_output['Flow Cell ID'] == i]['Top up'])
        # total_unique_outputs_per_flow_cell = len(unique_outputs_per_flow_cell)
        total_output_per_flow_cell = sum(flow_cells_to_output[flow_cells_to_output['Flow Cell ID'] == i]['Data output (Gb)'])
        output_per_flow_cell_df['Flow cell output (Gb)'][i] = total_output_per_flow_cell
        output_per_flow_cell_df['Top up'][i] = unique_topup_per_flow_cell
    # convert flow cell output total to numeric type for plotting
    output_per_flow_cell_df['Flow cell output (Gb)'] = pd.to_numeric(output_per_flow_cell_df['Flow cell output (Gb)'])
    # return flow_cells_per_experiment_df data frame
    return output_per_flow_cell_df
    
# get flow cells per experiment in two column list
def get_flow_cells_and_output_per_experiment(experiments, flow_cell_IDs, output):
    # take one column of experiments and one column of flow cell IDs from imported data frame as input
    # remove extraneous suffixes (e.g., "_topup") from experiment names
    # don't remove alphabetical characters altogether - e.g., PPMI_BLOOD_SSTEST is an experiment
    # experiment names should reflect independent brain isolates (e.g., PPMI_3080)
    experiments = experiments.str.replace(r'_topup', '', regex=True)
    experiments = experiments.str.replace(r'_recovery', '', regex=True)
    # dashes and underscores are the same thing so change all dashes to underscores
    experiments = experiments.str.replace(r'-', '_', regex=True)
    # make data frame of experiment names and flow cell IDs
    flow_cells_and_output_to_experiments = pd.concat([experiments, flow_cell_IDs, output], axis=1, join='inner')
    # find unique experiment names
    unique_experiments = np.unique(experiments)
    # initialized flow_cells_per_experiment list
    # create flow_cells_per_experiment_df data frame
    flow_cells_and_output_per_experiment_df = pd.DataFrame({'Experiment Name' : unique_experiments}, index=unique_experiments, columns=['Experiment Name','Flow Cells','Total output (Gb)'])
    # find flow cells and total output per unique experiment by iterating through unique experiment names
    for i in unique_experiments:
        unique_flow_cells_per_experiment = pd.unique(flow_cells_and_output_to_experiments[flow_cells_and_output_to_experiments['Experiment Name'] == i]['Flow Cell ID'])
        total_unique_flow_cells_per_experiment = len(unique_flow_cells_per_experiment)
        total_output_per_experiment = sum(flow_cells_and_output_to_experiments[flow_cells_and_output_to_experiments['Experiment Name'] == i]['Data output (Gb)'])
        flow_cells_and_output_per_experiment_df['Flow Cells'][i] = total_unique_flow_cells_per_experiment
        flow_cells_and_output_per_experiment_df['Total output (Gb)'][i] = total_output_per_experiment
    # convert output flow cell counts and total output totals to numeric type for plotting
    flow_cells_and_output_per_experiment_df['Flow Cells'] = pd.to_numeric(flow_cells_and_output_per_experiment_df['Flow Cells'])
    flow_cells_and_output_per_experiment_df['Total output (Gb)'] = pd.to_numeric(flow_cells_and_output_per_experiment_df['Total output (Gb)'])
    # return flow_cells_per_experiment_df data frame
    return flow_cells_and_output_per_experiment_df
    
# return flow cells per experiment distribution
# get total experiment count for each number of flow cells needed to complete experiment (approach 30x?)
def get_flow_cells_per_experiment_dist(column):
    # put distribution in flow_cells_per_experiment_dist
    # turn arange list into array and then 1d array with np.array() and flatten()
    flow_cells_per_experiment_dist = np.histogram(column, bins = np.array([np.arange(1,max(column)+2)]).flatten())
    # put output into labeled data frame for export
    flow_cells_per_experiment_dist_df = pd.DataFrame({'Flow Cells' : flow_cells_per_experiment_dist[1][0:max(column)], 'Frequency' : flow_cells_per_experiment_dist[0]}, columns=['Flow Cells', 'Frequency'])
    # return flow_cells_per_experiment_df data frame
    return flow_cells_per_experiment_dist_df

# get MinKNOW version distribution
def get_minknow_version_dist(column):
    # unique minknow versions
    unique_minknow_versions = np.unique(column)
    # create minknow_version_dist_df data frame
    minknow_version_dist_df = pd.DataFrame({'MinKNOW Version' : unique_minknow_versions}, index=unique_minknow_versions, columns=['MinKNOW Version', 'Frequency'])
    # count numbers of each in dataset
    for i in unique_minknow_versions:
        minknow_version_dist_df['Frequency'][i] = list(column).count(i)
    # return data frame with versions and counts per version
    return minknow_version_dist_df
    
# make summary statistic data frame
def make_summary_statistics_data_frame(summary_statistics_set, property_names):
    # set column names
    column_names = ['Property', 'Total', 'Min', 'Max', 'Mean', 'Median', 'Mode', 'Standard Deviation']
    # initialize data frame
    summary_statistics_df = pd.DataFrame(index=property_names, columns=column_names)
    # iterate through each property and populate data frame with summary_statistics_set attributes
    for idx, name in enumerate(property_names):
        summary_statistics_df.loc[name] = [property_names[idx],summary_statistics_set[idx].total,summary_statistics_set[idx].min,summary_statistics_set[idx].max,summary_statistics_set[idx].mean,summary_statistics_set[idx].median,summary_statistics_set[idx].mode,summary_statistics_set[idx].stdev]
    # return populated summary statistics data frame
    return summary_statistics_df
    
# identify topups and reconnections (flow cell moved and run restarted)
def identify_topups(column):
    # make output array "topups" as long as input column
    topups = [None] * len(column)
    for idx, i in enumerate(column):
        if "topup" in i:
            # if topup in sample name, set value to topup
            topups[idx] = "Top up"
        elif "recovery" in i:
            # if recovery in sample name, set value to recovery
            topups[idx] = "Recovery"
        elif "reconnected" in i:
            # if reconnected in sample name, set value to reconnection
            topups[idx] = "Reconnection"
        else:
            # if topup or other labels not in sample name, set value to "Initial run"
            topups[idx] = "Initial run"
    # return topups/no topups column
    return topups

# identify reconnections through sequence run flow cell ID (shared between runs), date, and experiment name
# reconnection if same name and flow cell ID as previous run; add value to topups column shown before
def identify_reconnections(data):
    data_with_reconnections = data.copy()
    # loop through rows of data frame on Flow Cell ID column
    for idx, i in enumerate(data['Flow Cell ID']):
        # see if flow cell ID in previous subset of list before current element
        if i in list(data['Flow Cell ID'][0:idx]):
            # get data frame for repeated flow cell
            duplicate_flow_cell_id_df = data[data['Flow Cell ID'] == i]
            # compare sample name for current run to all those with flow cell ID
            test_sample_name = data.loc[idx]["Sample Name"]
            # print(test_sample_name)
            # if same sample name more than once for given flow cell, name most recent run as reconnection
            if (sum(duplicate_flow_cell_id_df["Sample Name"] == test_sample_name) > 1):
                data_with_reconnections.loc[(idx,"Top up")] = "Reconnection"
    # return data frame with algorithmically detected reconnections in topups column
    return data_with_reconnections
    
# make violinplot/swarmplot figure worksheet in output workbook
def make_figure_worksheet(data,input_variable,workbook,worksheet_name,cutoff=None,title=None,top_up=None):
    # create worksheet for figure output
    worksheet=workbook.create_sheet(worksheet_name)
    # initialize raw data buffer for image
    imgdata=BytesIO()
    # initialize plot overall
    fig, ax = plt.subplots()
    # make swarm plot to show how data points overlap with distribution
    # color points based on top up variable (initial run, top up, reconnection, recovery)
    # replace color='black'
    if top_up is not None:
        rearranged_color_palette = [sb.color_palette()[0],sb.color_palette()[1],sb.color_palette()[4],sb.color_palette()[5]]
        ax = sb.swarmplot(data=data,x=input_variable,hue="Top up",hue_order=['Initial run','Top up','Reconnection','Recovery'],palette=rearranged_color_palette,edgecolor='white',linewidth=1)
    else:
        ax = sb.swarmplot(data=data,x=input_variable,color='black')
    # add violin plot using seaborn (sb.violinplot)
    # increase transparency to improve swarmplot visibility
    ax = sb.violinplot(data=data,x=input_variable,color='white',ax=ax)
    # add title if specified
    if title is not None:
        ax.set_title(title)
    # add red line for 90GB/30X cutoff or whatever necessary for specific plots
    # cutoff line defined by x value
    if cutoff is not None:
        ax.axvline(x=cutoff,color='red')
    # put figure in variable to prep for saving into buffer
    # fig = swarmplot.get_figure()
    # save figure as 200 dpi PNG into buffer
    fig.savefig(imgdata, format='png', dpi=200)
    # close figure
    fig.clf()
    # make openpyxl image from raw data
    img = openpyxl.drawing.image.Image(imgdata)
    # set location of image in worksheet (A1)
    img.anchor = 'A1'
    # add image to worksheet
    worksheet.add_image(img)
    
# make starting active pores vs. run data output scatterplot in output workbook
def make_active_pore_data_output_scatterplot(data,workbook,worksheet_name,title=None):
    # create worksheet for figure output
    worksheet=workbook.create_sheet(worksheet_name)
    # initialize raw data buffer for image
    imgdata=BytesIO()
    # initialize plot overall
    fig, ax = plt.subplots()
    # make scatterplot of active pores vs. per flow cell data output
    # include regression by using sb.regplot() function
    # had to remove regression to use hue keyword
    # color points by topup/not topup run
    rearranged_color_palette = [sb.color_palette()[0],sb.color_palette()[1],sb.color_palette()[4],sb.color_palette()[5]]
    ax = sb.scatterplot(data=data,x='Starting Active Pores',y='Data output (Gb)',hue="Top up",hue_order=['Initial run','Top up','Reconnection','Recovery'],palette=rearranged_color_palette)
    # add title if specified
    if title is not None:
        ax.set_title(title)
    # set minimum y and x to zero
    ax.set_xlim(left=0)
    ax.set_ylim(bottom=0)
    # add 90 GB/30x and 6500 pore cutoffs
    ax.axvline(x=6500,color='green')
    # also ONT warranty 5000 pore cutoff
    ax.axvline(x=5000,color='red')
    # add 90 GB/30x cutoff
    ax.axhline(y=90,color='gray')
    # put figure in variable to prep for saving into buffer
    # fig = swarmplot.get_figure()
    # save figure as 150 dpi PNG into buffer
    fig.savefig(imgdata, format='png', dpi=150)
    # close figure
    fig.clf()
    # make openpyxl image from raw data
    img = openpyxl.drawing.image.Image(imgdata)
    # set location of image in worksheet (A1)
    img.anchor = 'A1'
    # add image to worksheet
    worksheet.add_image(img)
    
# make starting active pores vs. flow cell data output scatterplot in output workbook
# this doesn't work because starting active pores are distinct per run
def make_active_pore_flow_cell_output_scatterplot(data,workbook,worksheet_name,title=None):     
    # create worksheet for figure output
    worksheet=workbook.create_sheet(worksheet_name)
    # initialize raw data buffer for image
    imgdata=BytesIO()
    # initialize plot overall
    fig, ax = plt.subplots()
    # make scatterplot of active pores vs. per flow cell data output
    # include regression by using sb.regplot() function
    # had to remove regression to use hue keyword
    # color points by topup/not topup run
    ax = sb.scatterplot(data=data,x='Starting Active Pores',y='Flow cell output (Gb)',hue="Top up",hue_order=['Initial run','Top up','Reconnection','Recovery'])
    # add title if specified
    if title is not None:
        ax.set_title(title)
    # set minimum y and x to zero
    ax.set_xlim(left=0)
    ax.set_ylim(bottom=0)
    # add 90 GB/30x and 6500 pore cutoffs
    ax.axvline(x=6500,color='green')
    # also ONT warranty 5000 pore cutoff
    ax.axvline(x=5000,color='red')
    # add 90 GB/30x cutoff
    ax.axhline(y=90,color='gray')
    # put figure in variable to prep for saving into buffer
    # fig = swarmplot.get_figure()
    # save figure as 150 dpi PNG into buffer
    fig.savefig(imgdata, format='png', dpi=150)
    # close figure
    fig.clf()
    # make openpyxl image from raw data
    img = openpyxl.drawing.image.Image(imgdata)
    # set location of image in worksheet (A1)
    img.anchor = 'A1'
    # add image to worksheet
    worksheet.add_image(img)
    
# make starting active pores vs. n50 scatterplot in output workbook
def make_active_pore_read_n50_scatterplot(data,workbook,worksheet_name,title=None):
    # create worksheet for figure output
    worksheet=workbook.create_sheet(worksheet_name)
    # initialize raw data buffer for image
    imgdata=BytesIO()
    # initialize plot overall
    fig, ax = plt.subplots()
    # make scatterplot of active pores vs. per flow cell data output
    # include regression by using sb.regplot() function
    # had to remove regression to use hue keyword
    # color points by topup/not topup run
    # use colors not used for cutoff lines
    rearranged_color_palette = [sb.color_palette()[0],sb.color_palette()[1],sb.color_palette()[4],sb.color_palette()[5]]
    ax = sb.scatterplot(data=data,x='Starting Active Pores',y='N50 (kb)',hue="Top up",hue_order=['Initial run','Top up','Reconnection','Recovery'],palette=rearranged_color_palette)
    # add title if specified
    if title is not None:
        ax.set_title(title)
    # set minimum y and x to zero
    ax.set_xlim(left=0)
    ax.set_ylim(bottom=0)
    # add 90 GB/30x and 6500 pore cutoffs
    ax.axvline(x=6500,color='green')
    # also ONT warranty 5000 pore cutoff
    ax.axvline(x=5000,color='red')
    # put figure in variable to prep for saving into buffer
    # fig = swarmplot.get_figure()
    # save figure as 150 dpi PNG into buffer
    fig.savefig(imgdata, format='png', dpi=150)
    # close figure
    fig.clf()
    # make openpyxl image from raw data
    img = openpyxl.drawing.image.Image(imgdata)
    # set location of image in worksheet (A1)
    img.anchor = 'A1'
    # add image to worksheet
    worksheet.add_image(img)
    
# make starting active pores vs. n50 scatterplot in output workbook
def make_read_n50_data_output_scatterplot(data,workbook,worksheet_name,title=None):
    # create worksheet for figure output
    worksheet=workbook.create_sheet(worksheet_name)
    # initialize raw data buffer for image
    imgdata=BytesIO()
    # initialize plot overall
    fig, ax = plt.subplots()
    # make scatterplot of active pores vs. per flow cell data output
    # include regression by using sb.regplot() function
    ax = sb.regplot(data=data,x='N50 (kb)',y='Data output (Gb)')
    # add title if specified
    if title is not None:
        ax.set_title(title)
    # set minimum y BUT not x to zero
    # no read N50 under 15 in example examined
    ax.set_ylim(bottom=0)
    # also ONT warranty 5000 pore cutoff
    ax.axhline(y=90,color='red')
    # put figure in variable to prep for saving into buffer
    # fig = swarmplot.get_figure()
    # save figure as 150 dpi PNG into buffer
    fig.savefig(imgdata, format='png', dpi=150)
    # close figure
    fig.clf()
    # make openpyxl image from raw data
    img = openpyxl.drawing.image.Image(imgdata)
    # set location of image in worksheet (A1)
    img.anchor = 'A1'
    # add image to worksheet
    worksheet.add_image(img)
    
# set up command line argument parser
parser = argparse.ArgumentParser(description='This program gets summary statistics from long read sequencing report data.')

# get input and output arguments
parser.add_argument('-input', action="store", dest="input_file", help="Input tab-delimited tsv file containing features extracted from long read sequencing reports.")
parser.add_argument('-output', action="store", dest="output_file", help="Output long read sequencing summary statistics XLSX")
parser.add_argument('-plot_title', action="store", default=None, dest="plot_title", help="Title for each plot in output XLSX (optional)")
# add boolean --plot_cutoff argument
parser.add_argument('--plot_cutoff', action=argparse.BooleanOptionalAction, default=True, dest="plot_cutoff", help="Include cutoff lines in violin plots (optional; default true; --no-plot_cutoff to override)")
# include failed run cutoff to exclude as well
parser.add_argument('-run_cutoff', action="store", default=1, dest="run_cutoff", help="Minimum data output per flow cell run to include (optional, 1 Gb default)")

# parse arguments
results = parser.parse_args()

# throw error if no input file provided
if results.input_file is None:
	quit('ERROR: No input file (-input) provided!')
        
# set default output filename
if results.output_file is None:
    results.output_file='output_summary_statistics.xlsx'

# read tab delimited output into pandas data frame
longread_extract_initial=pd.read_csv(results.input_file,sep='\t')

# use functions above
# first filter out low output runs
longread_extract = longread_extract_initial[longread_extract_initial['Data output (Gb)'] > results.run_cutoff]
# fix indices
longread_extract.reset_index(drop='True',inplace=True)
# add top up column to data frame
# avoid nested tuple warning
# longread_extract["Top up"] = identify_topups(longread_extract["Sample Name"])
# add after 12th column or last column (dataframe.shape[1])
longread_extract.insert(longread_extract.shape[1],"Top up",identify_topups(longread_extract["Sample Name"]),True)
# identify reconnections amongst flow cells
longread_extract = identify_reconnections(longread_extract)
# flow cells per experiment
longread_extract_flow_cells_and_output_per_experiment = get_flow_cells_and_output_per_experiment(longread_extract['Experiment Name'], longread_extract['Flow Cell ID'], longread_extract['Data output (Gb)'])
# output per flow cell
longread_extract_output_per_flow_cell = get_output_per_flow_cell(longread_extract['Flow Cell ID'], longread_extract['Data output (Gb)'], longread_extract['Top up'])
# flow cells per experiment distribution
longread_extract_flow_cells_per_experiment_dist = get_flow_cells_per_experiment_dist(longread_extract_flow_cells_and_output_per_experiment['Flow Cells'])
# summary statistics on...
# read N50
read_N50_summary_stats = get_summary_statistics(longread_extract['N50 (kb)'])
# sequence output
sequence_output_summary_stats = get_summary_statistics(longread_extract['Data output (Gb)'])
# starting active pores per run
starting_active_pores_summary_stats = get_summary_statistics(longread_extract['Starting Active Pores'])
# flow cells per experiment
flow_cells_per_experiment_summary_stats = get_summary_statistics(longread_extract_flow_cells_and_output_per_experiment['Flow Cells'])
# output per flow cell
output_per_flow_cell_summary_stats = get_summary_statistics(longread_extract_output_per_flow_cell['Flow cell output (Gb)'])
# total output per experiment
output_per_experiment_summary_stats = get_summary_statistics(longread_extract_flow_cells_and_output_per_experiment['Total output (Gb)'])

# combine summary stats into one list
combined_summary_stats = [read_N50_summary_stats,sequence_output_summary_stats,starting_active_pores_summary_stats,flow_cells_per_experiment_summary_stats,output_per_flow_cell_summary_stats,output_per_experiment_summary_stats]

# make data frame from combined_summary_stats
combined_property_names = ['Read N50 (kb)','Run data output (Gb)','Starting active pores','Flow cells per experiment','Flow cell output (Gb)', 'Total experiment output (Gb)']
combined_summary_stats_df = make_summary_statistics_data_frame(combined_summary_stats,combined_property_names)

# minknow version distribution
longread_extract_minknow_version_dist = get_minknow_version_dist(longread_extract['MinKNOW Version'])

# save data frames as tab-delimited file (.tsv)
# Example data structure
# Header 
# Property   Total   Min Max Mean    Median  Mode    Standard Deviation
# Read N50 (kb)
# Data output (Gb)
# Flow cells per experiment
# <empty>
# Flow Cells    Frequency
# 1 128
# 2 398
# etc.
# <empty>
# MinKNOW Version   Frequency
# 22.10.7   756
# 23.11.4   9
# etc.
# <empty>

# output data frames and figures to excel spreadsheet
writer = pd.ExcelWriter(results.output_file)
# write data frames with a row between each
# write combined summary stats
start_row = 0
combined_summary_stats_df.to_excel(writer, startrow=start_row, index=False, sheet_name='Summary statistics report')
# write flow cells per experiment distribution
# add 1 to row number after combined_summary_stats_df exported
start_row = start_row + len(combined_summary_stats_df) + 2
longread_extract_flow_cells_per_experiment_dist.to_excel(writer, startrow=start_row, index=False, sheet_name='Summary statistics report')
# write minknow version distribution
start_row = start_row + len(longread_extract_flow_cells_per_experiment_dist) + 2
longread_extract_minknow_version_dist.to_excel(writer, startrow=start_row, index=False, sheet_name='Summary statistics report')
# write flow cells and output per unique experiment on another worksheet
longread_extract_flow_cells_and_output_per_experiment.to_excel(writer, index=False, sheet_name='FC + output per experiment')
# close writer and save workbook
writer.close()
# then add svg figures
# use openpyxl and pipe image data into new worksheets
# append new worksheets to existing workbook
workbook = openpyxl.load_workbook(results.output_file)

# use make_figure_worksheet function to insert figures
# if/else depending on whether -plot_cutoff set
if results.plot_cutoff is True:
    # show topups in first three plots
    make_figure_worksheet(longread_extract,"N50 (kb)",workbook,'Read N50 plot',None,results.plot_title,True)
    make_figure_worksheet(longread_extract,"Data output (Gb)",workbook,'Run data output plot',90,results.plot_title,True)
    make_figure_worksheet(longread_extract,"Starting Active Pores",workbook,'Starting active pores plot',6500,results.plot_title,True)
    # now topups in next three plots
    make_figure_worksheet(longread_extract_flow_cells_and_output_per_experiment,"Flow Cells",workbook,'Flow cells per experiment plot',None,results.plot_title)
    make_figure_worksheet(longread_extract_flow_cells_and_output_per_experiment,"Total output (Gb)",workbook,'Output per experiment plot',90,results.plot_title)
    make_figure_worksheet(longread_extract_output_per_flow_cell,"Flow cell output (Gb)",workbook,'Output per flow cell plot',90,results.plot_title)
else:
    # show topups in first three plots
    make_figure_worksheet(longread_extract,"N50 (kb)",workbook,'Read N50 plot',None,results.plot_title,True)
    make_figure_worksheet(longread_extract,"Data output (Gb)",workbook,'Run data output plot',None,results.plot_title,True)
    make_figure_worksheet(longread_extract,"Starting Active Pores",workbook,'Starting active pores plot',None,results.plot_title,True)
    # now topups in next three plots
    make_figure_worksheet(longread_extract_flow_cells_and_output_per_experiment,"Flow Cells",workbook,'Flow cells per experiment plot',None,results.plot_title)
    make_figure_worksheet(longread_extract_flow_cells_and_output_per_experiment,"Total output (Gb)",workbook,'Output per experiment plot',None,results.plot_title)
    make_figure_worksheet(longread_extract_output_per_flow_cell,"Flow cell output (Gb)",workbook,'Output per flow cell plot',None,results.plot_title)
# use make_active_pore_data_output_scatterplot function to add starting active pore vs. data output scatterplot
make_active_pore_data_output_scatterplot(longread_extract,workbook,'Active pores vs. data output',results.plot_title)
# make_active_pore_flow_cell_output_scatterplot(longread_extract_output_per_flow_cell,workbook,'Active pores vs. flow cell output',results.plot_title)
make_active_pore_read_n50_scatterplot(longread_extract,workbook,'Active pores vs. read N50',results.plot_title)
make_read_n50_data_output_scatterplot(longread_extract,workbook,'Read N50 vs. data output',results.plot_title)
# save workbook when done
workbook.save(results.output_file)

# script complete
quit()